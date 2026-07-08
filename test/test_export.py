#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════╗
║              🧪 导出层单元测试 (test_export.py)                  ║
║                                                                  ║
║   【测试目标】                                                   ║
║   ① ExportManager 能导出装备库、科研期数和用户记录 CSV          ║
║   ② ExportManager 能导出欧非汇总 CSV                            ║
║   ③ ExportManager 能导出包含 CSV + Excel 的完整报告             ║
║                                                                  ║
║   【类比理解】                                                    ║
║   这组测试像"文件验收员"，确认导出的每份文件都真实可打开。      ║
╚══════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations

import csv
import os
import sys
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook


# ============================================================
# 📦 第一部分：导入依赖和准备路径
# ============================================================

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.data.export_manager import ExportManager, get_export_manager


# ============================================================
# 🧰 第二部分：pytest fixtures
# ============================================================

@pytest.fixture()
def export_manager() -> ExportManager:
    """返回全局导出管理器，保持和应用运行时一致的单例入口。"""
    return get_export_manager()


# ============================================================
# 🧪 第三部分：测试用例
# ============================================================

def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """读取 CSV 表头和数据行，方便测试分别验证结构与内容。"""
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def export_csv(export_manager: ExportManager, kind: str, output_path: Path) -> Path:
    """按测试参数选择对应导出方法，避免测试中重复分支细节。"""
    if kind == "equipment":
        return Path(export_manager.export_equipment_library(str(output_path)))
    if kind == "research":
        return Path(export_manager.export_research_phases(str(output_path)))
    if kind == "today":
        return Path(export_manager.export_user_records(output_path=str(output_path)))
    if kind == "all":
        return Path(export_manager.export_all_user_records(str(output_path)))
    if kind == "luck":
        return Path(export_manager.export_luck_summary(str(output_path)))
    raise ValueError(f"未知导出类型: {kind}")


@pytest.mark.parametrize(
    ("kind", "filename", "expected_fields", "minimum_rows"),
    [
        ("equipment", "equipment_library.csv", {"equipment_id", "name", "rarity_name", "image_path"}, 14),
        ("research", "research_phases.csv", {"phase_number", "name", "equipment_details"}, 6),
        ("today", "today_records.csv", {"date", "equipment_id", "equipment_count", "fragment_count"}, 0),
        ("all", "all_records.csv", {"date", "equipment_id", "equipment_count", "fragment_count"}, 0),
        ("luck", "luck_summary.csv", {"scope", "phase_number", "luck_value", "luck_level"}, 1),
    ],
)
def test_export_manager_writes_csv(
    export_manager: ExportManager,
    tmp_path: Path,
    kind: str,
    filename: str,
    expected_fields: set[str],
    minimum_rows: int,
) -> None:
    """五类 CSV 导出都应生成文件，并包含约定表头。"""
    output_path = tmp_path / filename
    exported_path = export_csv(export_manager, kind, output_path)
    fieldnames, rows = read_csv(exported_path)

    assert exported_path.exists()
    assert exported_path.stat().st_size > 0
    assert expected_fields.issubset(set(fieldnames))
    assert len(rows) >= minimum_rows


def test_export_manager_writes_full_report(export_manager: ExportManager, tmp_path: Path) -> None:
    """完整报告应包含五份 CSV 和一个带固定工作表的 Excel 文件。"""
    result: dict[str, Any] = export_manager.export_full_report(str(tmp_path / "full_report"))
    output_dir = Path(result["output_dir"])
    workbook_path = Path(result["workbook"])

    assert output_dir.exists()
    assert workbook_path.exists()
    assert (output_dir / "equipment_library.csv").exists()
    assert (output_dir / "research_phases.csv").exists()
    assert any(path.name.startswith("today_") and path.suffix == ".csv" for path in output_dir.iterdir())
    assert any(path.name.startswith("all_records_") and path.suffix == ".csv" for path in output_dir.iterdir())
    assert any(path.name.startswith("luck_summary_") and path.suffix == ".csv" for path in output_dir.iterdir())

    workbook = load_workbook(workbook_path)
    expected_sheets = {"Equipment Library", "Research Phases", "Today Records", "All Records", "Luck Summary"}
    assert expected_sheets.issubset(set(workbook.sheetnames))
    workbook.close()


def test_export_user_records_rejects_invalid_date(export_manager: ExportManager, tmp_path: Path) -> None:
    """ExportManager 应拒绝非法日期，避免后续 GUI 绕过 CLI 校验。"""
    with pytest.raises(ValueError, match="YYYY-MM-DD"):
        export_manager.export_user_records("2026-99-99", str(tmp_path / "bad_date.csv"))
