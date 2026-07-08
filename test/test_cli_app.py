#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════╗
║               🧪 CLI 入口回归测试 (test_cli_app.py)             ║
║                                                                  ║
║   【测试目标】                                                   ║
║   ① main.py 只保留启动壳，不再堆放 CLI 逻辑                      ║
║   ② core/cli/app.py 承载 status / record / export               ║
║   ③ CLI 可以导出 CSV 和 Excel 报告                               ║
║                                                                  ║
║   【类比理解】                                                    ║
║   这组测试像"入口安检员"，确认门口很清爽，真正办事的人在里面。  ║
╚══════════════════════════════════════════════════════════════════╝
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path
from typing import Callable, Sequence

import pytest
from openpyxl import load_workbook


# ============================================================
# 📦 第一部分：导入依赖和准备路径
# ============================================================

ROOT = Path(__file__).resolve().parents[1]
MAIN_PATH = ROOT / "main.py"
CLI_APP_PATH = ROOT / "core" / "cli" / "app.py"


# ============================================================
# 🧰 第二部分：pytest fixtures
# ============================================================

@pytest.fixture()
def cli_env() -> dict[str, str]:
    """为子进程准备 UTF-8 环境，避免中文输出在 Windows 终端乱码。"""
    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    return env


@pytest.fixture()
def run_cli(cli_env: dict[str, str]) -> Callable[[Sequence[str]], subprocess.CompletedProcess[str]]:
    """返回一个 CLI 运行器，让每个测试都用同一套执行方式。"""

    def _run(args: Sequence[str]) -> subprocess.CompletedProcess[str]:
        """运行 main.py 子命令，并把 stdin 关闭成非交互模式。"""
        return subprocess.run(
            [sys.executable, *args],
            cwd=str(ROOT),
            env=cli_env,
            input="",
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )

    return _run


# ============================================================
# 🧪 第三部分：测试用例
# ============================================================

def read_text(path: Path) -> str:
    """读取 UTF-8 文本，供结构检查类测试复用。"""
    return path.read_text(encoding="utf-8")


def test_main_py_is_thin_startup_wrapper() -> None:
    """main.py 应只负责准备路径并转交给 core.cli.app.main。"""
    source = read_text(MAIN_PATH)

    assert "from core.cli.app import main as cli_main" in source
    assert "PROJECT_ROOT" in source
    assert "sys.path.insert" in source
    assert "def command_status" not in source
    assert "def command_record" not in source
    assert "def command_export" not in source
    assert "def build_parser" not in source
    assert "def interactive_menu" not in source


def test_cli_app_owns_command_logic() -> None:
    """core.cli.app 应集中承载三个核心命令和参数解析器。"""
    source = read_text(CLI_APP_PATH)

    assert "def command_status" in source
    assert "def command_record" in source
    assert "def command_export" in source
    assert "def build_parser" in source
    assert "def main(" in source
    assert 'if __name__ == "__main__"' not in source


@pytest.mark.parametrize("command_name", ["status", "record", "export"])
def test_cli_help_lists_core_commands(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    command_name: str,
) -> None:
    """--help 应展示 v0.4.0 只保留的三个核心子命令。"""
    result = run_cli(["main.py", "--help"])

    assert result.returncode == 0
    assert command_name in result.stdout


def test_status_command_runs_with_trend_phase(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
) -> None:
    """status 应能输出当前状态，并支持指定历史趋势期数。"""
    result = run_cli(["main.py", "status", "--trend-phase", "1"])

    assert result.returncode == 0
    assert "v0.4.0" in result.stdout
    assert "status" in result.stdout
    assert "今天的总览" in result.stdout


def test_record_command_exits_cleanly_without_interactive_stdin(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
) -> None:
    """record 在非交互测试环境里应直接提示并安全退出。"""
    result = run_cli(["main.py", "record"])

    assert result.returncode == 0
    assert "record" in result.stdout
    assert "交互式终端" in result.stdout


def test_record_command_accepts_dry_run_set(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
) -> None:
    """record --set --dry-run 应能校验单条记录且不写入文件。"""
    result = run_cli(["main.py", "record", "--set", "S1-001", "2", "30", "--dry-run"])

    assert result.returncode == 0
    assert "校验通过" in result.stdout
    assert "未写入数据" in result.stdout


def test_record_command_rejects_unknown_equipment(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
) -> None:
    """非交互录入时，未知装备 ID 应被拦截。"""
    result = run_cli(["main.py", "record", "--set", "NO-SUCH-ID", "1", "1", "--dry-run"])

    assert result.returncode == 1
    assert "装备 ID 不存在" in result.stdout


def test_record_command_accepts_batch_file_dry_run(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
) -> None:
    """record --batch-file --dry-run 应能校验标准 CSV 批量输入。"""
    batch_file = tmp_path / "records.csv"
    batch_file.write_text(
        "equipment_id,equipment_count,fragment_count\n"
        "S1-001,2,30\n"
        "S1-002,1,10\n",
        encoding="utf-8-sig",
    )

    result = run_cli(["main.py", "record", "--batch-file", str(batch_file), "--dry-run"])

    assert result.returncode == 0
    assert "校验通过" in result.stdout
    assert "2 条记录" in result.stdout


def test_record_command_rejects_bad_batch_file(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
) -> None:
    """批量 CSV 缺少必要字段时应返回失败。"""
    batch_file = tmp_path / "bad_records.csv"
    batch_file.write_text("equipment_id,equipment_count\nS1-001,2\n", encoding="utf-8-sig")

    result = run_cli(["main.py", "record", "--batch-file", str(batch_file), "--dry-run"])

    assert result.returncode == 1
    assert "缺少字段" in result.stdout


@pytest.mark.parametrize(
    ("kind", "filename"),
    [
        ("equipment", "equipment.csv"),
        ("research", "research.csv"),
        ("today", "today.csv"),
        ("all", "all.csv"),
        ("luck", "luck.csv"),
    ],
)
def test_export_command_writes_csv_files(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
    kind: str,
    filename: str,
) -> None:
    """export 的五种 CSV 类型都应生成非空文件。"""
    output_file = tmp_path / filename
    result = run_cli(["main.py", "export", "--kind", kind, "--output", str(output_file)])

    assert result.returncode == 0
    assert output_file.exists()
    assert output_file.stat().st_size > 0


@pytest.mark.parametrize("kind", ["equipment", "research", "today", "all", "luck"])
def test_export_command_respects_output_dir(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
    kind: str,
) -> None:
    """普通导出类型使用 --output-dir 时应写入指定目录，而不是默认目录。"""
    output_dir = tmp_path / f"{kind}_export_dir"
    result = run_cli(["main.py", "export", "--kind", kind, "--output-dir", str(output_dir)])

    assert result.returncode == 0
    assert output_dir.exists()
    csv_files = list(output_dir.glob("*.csv"))
    assert len(csv_files) == 1
    assert csv_files[0].stat().st_size > 0


def test_export_command_writes_full_report(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
) -> None:
    """完整报告应同时生成多份 CSV 和一个 Excel 工作簿。"""
    output_dir = tmp_path / "full_report"
    result = run_cli(["main.py", "export", "--kind", "full", "--output-dir", str(output_dir)])

    assert result.returncode == 0
    assert output_dir.exists()
    assert (output_dir / "equipment_library.csv").exists()
    assert (output_dir / "research_phases.csv").exists()
    assert any(path.name.startswith("today_") for path in output_dir.iterdir())
    assert any(path.name.startswith("all_records_") for path in output_dir.iterdir())
    assert any(path.name.startswith("luck_summary_") for path in output_dir.iterdir())

    workbook_candidates = sorted(output_dir.glob("full_report_*.xlsx"))
    assert len(workbook_candidates) == 1

    workbook = load_workbook(workbook_candidates[0])
    expected_sheets = {"Equipment Library", "Research Phases", "Today Records", "All Records", "Luck Summary"}
    assert expected_sheets.issubset(set(workbook.sheetnames))
    workbook.close()


def test_export_full_report_accepts_output_as_directory(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
) -> None:
    """完整报告也应接受 --output 作为输出目录，减少用户记忆负担。"""
    output_dir = tmp_path / "full_report_by_output"
    result = run_cli(["main.py", "export", "--kind", "full", "--output", str(output_dir)])

    assert result.returncode == 0
    assert output_dir.exists()
    assert (output_dir / "equipment_library.csv").exists()
    assert len(sorted(output_dir.glob("full_report_*.xlsx"))) == 1


def test_export_today_rejects_invalid_date(
    run_cli: Callable[[Sequence[str]], subprocess.CompletedProcess[str]],
    tmp_path: Path,
) -> None:
    """非法日期应被 argparse 拦截，不应生成导出文件。"""
    output_file = tmp_path / "bad_date.csv"
    result = run_cli(["main.py", "export", "--kind", "today", "--date", "2026-99-99", "--output", str(output_file)])

    assert result.returncode == 2
    assert "YYYY-MM-DD" in result.stderr
    assert not output_file.exists()
